"""
Shared on-demand format conversion: query one country's rows out of the master
parquet, write it as csv/xlsx/parquet, cache the result on disk.

Used by both scripts/build_catalog.py (dev-time -- to record accurate file sizes in
catalog.json and pre-warm the cache) and server/main.py (runtime -- generate on first
request, serve the cached file on every request after). Same code path both places on
purpose: a size recorded in catalog.json should be the exact byte count a user's
download gets, not an estimate.

Deploying without the full ~1.8 GB `to JD/pipeline/data/clean/` tree depends on this:
the deploy host only needs the single master parquet (80 MB, synced from Google Drive
by sync_from_drive.py) -- csv/xlsx are never stored at rest, only ever generated from
it, exactly reproducing what write_output.py / export_xlsx.py already do.
"""

from __future__ import annotations

import csv
from pathlib import Path

import duckdb
import pandas as pd
from openpyxl.utils import get_column_letter

# Mirrors pipeline/scripts/lib/schema.py:TEXT_SAFE_COLUMNS. Duplicated, not
# imported: the canonical schema is a frozen historical fact (these are exactly the
# 24 delivered columns, unchanged since the pipeline's second pass), and the deploy
# host must not need `to JD/` on its Python path just to read this one constant.
TEXT_SAFE_COLUMNS = {
    "country_iso2", "country_iso3", "country_en", "country_lc",
    "continent_code", "continent_name",
    "postal_code", "postal_code_norm", "place_name",
    "region_1_lc", "region_1_en", "region_1_iso_code",
    "region_2_lc", "region_2_en",
    "region_3_lc", "region_3_en", "region_4_lc", "region_4_en",
    "region_5_lc", "region_5_en",
    "timezone", "utc_offset",
}

# Excel's hard ceiling is 1,048,576 rows including the header (see export_xlsx.py).
# Nothing in the current dataset gets close (CA is the largest at 892,800), but a
# future country could -- the guard matches the pipeline's own behavior: skip, don't
# truncate silently.
EXCEL_ROW_LIMIT = 1_048_575

TEXT_FORMAT = "@"  # Excel's "Text" number format

# How many rows a "sample" CSV keeps. Small enough to commit directly to git for
# every country regardless of its full size (CA's full CSV is 157 MB -- past
# GitHub's 100 MB hard per-file limit; its 100-row sample is a few KB) -- used by
# build_catalog.py for the "select several, download as one zip" bulk feature and
# as a quick, always-available look at a country's columns and shape. The full
# file, when needed, is a Drive link instead (see link_drive_files.py).
SAMPLE_ROWS = 100

# Every country's data is inherently many-to-many: one postal code can cover many
# admin areas and vice versa (e.g. DE's 23,296 rows collapse to 652 distinct
# 3-level region tuples -- confirmed generic, proven out first on Australia, now
# applied to every covered country). These are the two alternate groupings a
# country can be deduped by.
POSTAL_CODE_KEY = ["postal_code"]
ADMIN_AREA_KEY = [
    "region_1_lc", "region_1_en", "region_2_lc", "region_2_en",
    "region_3_lc", "region_3_en", "region_4_lc", "region_4_en",
    "region_5_lc", "region_5_en",
]
VIEW_KEYS = {"postal_codes": POSTAL_CODE_KEY, "admin_areas": ADMIN_AREA_KEY}


def dedupe(df: pd.DataFrame, key_cols: list[str]) -> pd.DataFrame:
    """One row per distinct key_cols combination (e.g. one row per postal_code, or
    one row per admin-area tuple). Every other column keeps its value only if every
    row in the group agrees; otherwise it's blanked -- a postal code spanning 111
    localities (a real AU example: postcode 0822) should not silently show just one
    of them as if it were the only one.

    Blank (NaN for numeric, "" for text) is treated as "no opinion", not as a vote
    for blank -- a postal code with 110 rows naming a locality and 1 row missing it
    should still show that locality, not blank it out because of the 1 hole. This
    matches how the rest of this dataset already treats missing data (see
    build_catalog.py's nonblank()): honest and consistent, not a new convention.

    Built from two vectorized (C-level) groupby methods -- first() for a
    representative value per column, nunique() to detect disagreement -- rather
    than a custom Python callable passed to .agg(). That earlier version was
    correct but never finished on CA: a callable-per-group-per-column doesn't
    vectorize, so its cost scales with group count x column count in Python-level
    function calls. AU's ~3-18K groups hid this completely (seconds); CA's
    ~892K postal-code groups didn't finish in 48+ minutes before being killed and
    rewritten to this. Confirmed both versions agree on AU's known cases (postcode
    0822, the cross-border 0872, the central-Sydney admin tuple) before switching.

    dropna=False is required, not optional: region_4/region_5 are entirely NaN for
    most countries (confirmed: AU's region_4_lc is 100% NaN, not empty string), and
    pandas' groupby silently DROPS every row when any key column is NaN unless told
    not to -- with the default, an admin-areas dedupe would return zero rows for
    exactly the countries this feature is built for.
    """
    value_cols = [c for c in df.columns if c not in key_cols]

    # "" and NaN both mean "no opinion" for a text column -- normalize once up
    # front so first()/nunique() below (both dropna=True) treat them identically,
    # without re-deriving this per column inside a slow per-group callable.
    blanked = df.copy()
    for c in value_cols:
        if not pd.api.types.is_numeric_dtype(df[c]):
            blanked[c] = df[c].where(df[c].astype(str).str.strip() != "")

    g = blanked.groupby(key_cols, dropna=False, sort=False)
    # Independently per column: that column's first non-null value in the group --
    # "assume this until disagreement is found" (skipna=True is first()'s default).
    result = g[value_cols].first()
    disagreement = g[value_cols].nunique(dropna=True)

    for c in value_cols:
        blank = pd.NA if pd.api.types.is_numeric_dtype(df[c]) else ""
        result.loc[disagreement[c] > 1, c] = blank

    result = result.reset_index()
    return result[list(df.columns)]  # restore canonical column order


def query_country(duck: duckdb.DuckDBPyConnection, master_parquet: Path, iso2: str) -> pd.DataFrame:
    """One country's full row set, straight off the master parquet.

    country_iso2 is pushed into the parquet scan (DuckDB predicate pushdown), so this
    reads only that country's row groups -- not the full 3.66M-row file -- exactly
    like the /preview endpoint already does.
    """
    return duck.execute(
        "SELECT * FROM read_parquet($p) WHERE country_iso2 = $iso2",
        {"p": str(master_parquet), "iso2": iso2},
    ).df()


def write_csv(df: pd.DataFrame, out_path: Path) -> None:
    # QUOTE_NONNUMERIC quotes every text-safe column as a literal string -- the same
    # mechanism write_output.py uses. Reading this file back with dtype=str (or via
    # the XLSX below) keeps leading zeros; a bare `pd.read_csv` still won't, which is
    # the CSV-vs-Excel warning the UI repeats.
    df.to_csv(out_path, index=False, quoting=csv.QUOTE_NONNUMERIC, na_rep="")


def write_xlsx(df: pd.DataFrame, out_path: Path, sheet_name: str) -> bool:
    """Returns False (writing nothing) if the row count exceeds Excel's limit --
    mirrors export_xlsx.py's own skip-and-report behavior rather than truncating."""
    if len(df) > EXCEL_ROW_LIMIT:
        return False
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        for i, col in enumerate(df.columns, start=1):
            if col in TEXT_SAFE_COLUMNS:
                letter = get_column_letter(i)
                # Stamp the whole column, including the header cell, so Excel treats
                # every value as text -- this is the actual fix for the leading-zero
                # bug; a plain df.to_excel() alone would reintroduce it.
                for cell in ws[letter]:
                    cell.number_format = TEXT_FORMAT
        ws.freeze_panes = "A2"
    return True


def write_parquet(df: pd.DataFrame, out_path: Path) -> None:
    df.to_parquet(out_path, index=False, compression="zstd")


def materialize(
    duck: duckdb.DuckDBPyConnection,
    master_parquet: Path,
    cache_dir: Path,
    iso2: str,
    fmt: str,
    view: str = "all_rows",
) -> Path | None:
    """Return the cached file for (iso2, view, fmt), generating it first if needed.

    view "all_rows" (the default) is every row, unchanged from before this
    parameter existed -- every existing caller keeps working with no code changes.
    "postal_codes"/"admin_areas" additionally dedupe via dedupe() before writing.

    Returns None if the country has no rows, or (xlsx only) exceeds Excel's row
    limit -- callers should treat that the same as "format not available".
    """
    if fmt not in ("csv", "xlsx", "parquet"):
        raise ValueError(f"unknown format: {fmt}")
    if view != "all_rows" and view not in VIEW_KEYS:
        raise ValueError(f"unknown view: {view}")

    # Cache filename only grows a `.view` segment for the two new views, so every
    # existing cache entry (every country, every format, from before this feature)
    # stays exactly where it was -- nothing is invalidated by adding this parameter.
    out_path = cache_dir / (f"{iso2}.{fmt}" if view == "all_rows" else f"{iso2}.{view}.{fmt}")
    # Staleness, not just existence: a country's canonical data can change (a new
    # 數據更新 revision lands) without app/on_demand_cache/ being cleared. Without
    # this mtime check, materialize() would happily keep serving a country's *old*
    # csv/xlsx/parquet forever after such an update -- exactly what would have
    # happened silently to AU's cache after its 2026-08-13 source swap.
    if out_path.exists() and out_path.stat().st_mtime >= master_parquet.stat().st_mtime:
        return out_path

    df = query_country(duck, master_parquet, iso2)
    if df.empty:
        return None
    if view != "all_rows":
        df = dedupe(df, VIEW_KEYS[view])

    cache_dir.mkdir(parents=True, exist_ok=True)
    tmp_path = out_path.with_suffix(out_path.suffix + ".tmp")
    try:
        if fmt == "csv":
            write_csv(df, tmp_path)
        elif fmt == "xlsx":
            if not write_xlsx(df, tmp_path, iso2):
                return None
        elif fmt == "parquet":
            write_parquet(df, tmp_path)
        tmp_path.rename(out_path)  # atomic -- a concurrent reader never sees a partial file
    finally:
        tmp_path.unlink(missing_ok=True)

    return out_path
